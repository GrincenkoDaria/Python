from asyncore import read
import openpyxl
import json
book = openpyxl.open("Knigi.xlsx", read_only = True)
sheet = book.active
cells = sheet["B1":"C11"]
print(sheet["B2".value])

for name, year, in cells:
  print(name.value,year.value)
  
for i in range(1,sheet.max_row +1):
  author = sheet[i][0].value
  name = sheet[i][1].value
  year = sheet[i][2].value
  rating = sheet[i][3].value
  print(i, author, name,year,rating)
  
  for i in sheet.iter_rows(min_row=2, max_row=20, min_col=1,max_col=3):
    for cell in row:
      print(cell.value, end=" ")
    print()

sheets = book.worhsheets[2]
print(sheet_2['A2'])
 
book = openpyxl.Workbook()
sheet = book.active
sheet["B2"] =100
sheet["C11"]= 'hello'
sheet [1][0].value= 'world'

book.save("my_book.xlsx")
book.close()

with open ('movies.json') as file:
  data = json.load(file)
  
for movie in data["movies"]:
  id = movie["id"]
  title = movie["title"]
  year = movie["year"]
  genres = movie["genres"]
  director = movie["director"]
  actors = movie["actors"]
  print(id,title, year, genres, director,actors)